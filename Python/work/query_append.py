from openpyxl import load_workbook
import openpyxl
import pandas as pd

inputfile = r'C:\svn_home\个人目录\李孟南（）\2222.xlsx'
outputfile = r'C:\Users\sas\Desktop\查询结果.xlsx'


#   写入已存在的xlsx文件第一种方法
class Write_excel(object):
    def __init__(self, filename):
         self.filename =  outputfile
         self.wb = load_workbook(self.filename)
         self.ws = self.wb.active  # 激活sheet
    def write(self, row_n, value):
         self.ws.cell(row_n,value )
         self.wb.save(self.filename)

def main():
    data = pd.DataFrame(pd.read_excel(inputfile))
    print(data.columns.values)
    row = input("请输入行名称:")
    information = input("请输入查询内容:")
    result = str(data.loc[data[row] == information])
    print(result)
    if row in data:
        print("查询成功")
        print('请输入插入位置:')
        a = int(input('请输入行:'))
        b = int(input('请输入列:'))
        we = Write_excel(r'C:\Users\sas\Desktop\查询结果.xlsx')
        we.write(a,result)
    else:
        print("输入信息有误")

main()

